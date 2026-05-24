import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

file_path = r"D:\Kê hoạch sản xuât\FORECAST\W21.(18-23-05-) SALEFORECAST 2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

ws = None
for name in wb.sheetnames:
    if "W21" in name or "18-23-05" in name:
        ws = wb[name]
        break
if ws is None:
    ws = wb.active

print(f"Reading rows 250 to 350 in sheet: {ws.title}")

for r_idx in range(250, 350):
    row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 10)]
    # if any value in the row is not None:
    if any(val is not None for val in row_vals):
        print(f"Row {r_idx:3d}: {row_vals}")

wb.close()
