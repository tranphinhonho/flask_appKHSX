import openpyxl
import os

with open("scratch_check_khsx.txt", "w", encoding="utf-8") as f:
    file_path = r"D:\Kê hoạch sản xuât\laptrinh vao\output\KHSX_18-05-2026.xlsx"
    f.write(f"Reading file: {file_path}\n")
    if not os.path.exists(file_path):
        f.write("File does not exist!\n")
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # The sheet name is result.date, which is "18-05-2026"
        sheet_name = "18-05-2026"
        if sheet_name not in wb.sheetnames:
            f.write(f"Sheet {sheet_name} not found! Available sheets: {wb.sheetnames}\n")
            ws = wb.active
        else:
            ws = wb[sheet_name]
            
        f.write(f"Sheet title: {ws.title}\n")
        f.write("\nRows 7 to 41:\n")
        f.write(f"{'Row':<5} | {'STT':<5} | {'TEN CAM':<15} | {'ME':<5} | {'TONG TAN':<10} | {'STAR 25':<8} | {'CP 25':<8} | {'SILO TRUCK':<10} | {'LINE CV':<8} | {'LINE PK':<8}\n")
        f.write("-" * 100 + "\n")
        for r in range(7, 42):
            stt = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            batches = ws.cell(row=r, column=3).value
            tons = ws.cell(row=r, column=4).value
            star_25 = ws.cell(row=r, column=9).value # STAR 25 is column I (column 9)
            cp_25 = ws.cell(row=r, column=7).value # CP 25 is column G (column 7)
            silo = ws.cell(row=r, column=20).value # Silo Truck is column T (column 20)
            line_cv = ws.cell(row=r, column=22).value # Line CV is V (column 22)
            line_pk = ws.cell(row=r, column=23).value # Line PK is W (column 23)
            f.write(f"{r:<5} | {str(stt):<5} | {str(name):<15} | {str(batches):<5} | {str(tons):<10} | {str(star_25):<8} | {str(cp_25):<8} | {str(silo):<10} | {str(line_cv):<8} | {str(line_pk):<8}\n")
