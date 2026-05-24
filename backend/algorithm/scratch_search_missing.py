import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

file_path = r"D:\Kê hoạch sản xuât\FORECAST\W21.(18-23-05-) SALEFORECAST 2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

ws = None
for name in wb.sheetnames:
    if "W21" in name or "18-23-05" in name:
        ws = wb[name]
        break
if ws is None:
    ws = wb.active

print(f"Reading sheet: {ws.title}")

targets = ['502', '562P', 'BS07TA', 'BS09TA']

print(f"{'FORMULAR':<10} | {'DIE':<5} | {'PACK':<6} | {'HIGRO':<10} | {'CP':<10} | {'STAR':<10} | {'NUVO':<10} | {'NASA':<10} | {'FARM':<10}")
print("-" * 110)

for row in ws.iter_rows(min_row=8, max_row=260, max_col=10):
    vals = [str(cell.value).strip().upper() if cell.value is not None else 'NONE' for cell in row[:9]]
    found = False
    for t in targets:
        # Check if t is exact match or part of word
        for v in vals:
            if t.upper() in v:
                found = True
                break
    if found:
        print(" | ".join(f"{v:<10}" for v in vals))

wb.close()
