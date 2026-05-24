import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

file_path = r"D:\Kê hoạch sản xuât\FORECAST\W21.(18-23-05-) SALEFORECAST 2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

ws = None
for name in wb.sheetnames:
    if "W21" in name or "18-23-05" in name:
        ws = wb[name]
        break
if ws is None:
    ws = wb.active

print(f"Reading sheet: {ws.title}")

targets = ['502', 'BS07TA', 'BS09TA']

for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    for c_idx, cell in enumerate(row, start=1):
        if cell is not None:
            s_val = str(cell).strip().upper()
            for t in targets:
                if t.upper() in s_val:
                    print(f"Found {t} at Row {r_idx}, Col {c_idx}: {cell}")

wb.close()
