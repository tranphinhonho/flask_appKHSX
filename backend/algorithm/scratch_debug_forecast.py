import openpyxl

file_path = r"D:\Kê hoạch sản xuât\FORECAST\W21.(18-23-05-) SALEFORECAST 2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

output_path = r"D:\Kê hoạch sản xuât\laptrinh vao\scratch_debug_forecast.txt"
with open(output_path, "w", encoding="utf-8") as f:
    f.write(f"Workbook sheets: {wb.sheetnames}\n")
    ws = wb.active
    f.write(f"Active sheet: {ws.title}\n")
    
    # Check sheet named 'W21.18-23-05-2026' or similar
    target_sheet = None
    for name in wb.sheetnames:
        if "W21" in name:
            target_sheet = name
            break
            
    if target_sheet:
        ws = wb[target_sheet]
        f.write(f"Using found sheet: {target_sheet}\n")
    else:
        f.write("Sheet with 'W21' not found, using active sheet\n")
        
    f.write("Rows 250 to 300 details:\n")
    for r in range(250, 301):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        f.write(f"Row {r:3d}: {row_vals}\n")

wb.close()
print("Debug finished")
