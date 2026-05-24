import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

file_path = r"D:\Kê hoạch sản xuât\FORECAST\W21.(18-23-05-) SALEFORECAST 2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

# print sheet names
print("Sheets in forecast:", wb.sheetnames)

# We find the latest or active sheet
ws = None
for name in wb.sheetnames:
    if "W21" in name or "18-23-05" in name:
        ws = wb[name]
        break
if ws is None:
    ws = wb.active

print(f"Reading sheet: {ws.title}")

# Print mapping
print(f"{'FORMULAR':<10} | {'DIE':<5} | {'PACK':<6} | {'HIGRO':<10} | {'CP':<10} | {'STAR':<10} | {'NUVO':<10} | {'NASA':<10} | {'FARM':<10}")
print("-" * 110)

count = 0
for row in ws.iter_rows(min_row=8, max_row=200, max_col=10):
    formular = row[0].value
    if formular is None:
        continue
    
    die = row[1].value
    pack = row[2].value
    higro = row[3].value
    cp = row[4].value
    star = row[5].value
    nuvo = row[6].value
    nasa = row[7].value
    farm = row[8].value
    
    print(f"{str(formular):<10} | {str(die):<5} | {str(pack):<6} | {str(higro):<10} | {str(cp):<10} | {str(star):<10} | {str(nuvo):<10} | {str(nasa):<10} | {str(farm):<10}")
    count += 1

wb.close()
