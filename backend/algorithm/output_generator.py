"""
output_generator.py - Xuất kế hoạch sản xuất ra file Excel chất lượng cao dựa trên template
"""
import os
import sys
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import KHSXResult, DemandItem
import config

# Fix encoding cho Windows
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass


def generate_khsx_excel(result, output_dir, filename=None):
    """
    Xuất KHSX ra file Excel theo format nhà máy bằng cách sử dụng file mẫu (KHSX_FILE).
    
    Args:
        result: KHSXResult
        output_dir: Thư mục output
        filename: Tên file (None = auto)
    
    Returns:
        str: Đường dẫn file đã tạo
    """
    if filename is None:
        filename = f"KHSX_{result.date}.xlsx"
    
    filepath = os.path.join(output_dir, filename)
    
    # 1. Load workbook template (KHSX_FILE)
    print(f"  📖 Đang đọc file mẫu: {config.KHSX_FILE}...")
    wb = openpyxl.load_workbook(config.KHSX_FILE, data_only=False)
    
    # 2. Sao chép sheet '19' làm template cho sheet ngày mới
    if '19' in wb.sheetnames:
        template_sheet_name = '19'
    else:
        # Fallback nếu không thấy sheet 19, lấy sheet đầu tiên
        template_sheet_name = [s for s in wb.sheetnames if s.isdigit()][0]
        
    print(f"  📋 Sử dụng sheet '{template_sheet_name}' làm layout mẫu...")
    template_sheet = wb[template_sheet_name]
    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = result.date
    
    # 3. Di chuyển sheet mới lên vị trí đầu tiên (index 0)
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(new_sheet)))
    
    # 4. Định dạng và điền ngày vào cell U3
    # result.date dạng DD-MM-YYYY, đổi sang Ngày:…DD…/……MM….. /…YYYY
    parts = result.date.split('-')
    if len(parts) == 3:
        day = parts[0]
        month = str(int(parts[1]))  # Bỏ số 0 ở đầu nếu có
        year = parts[2]
        date_formatted = f"Ngày:…{day}…/……{month}…../…{year}"
    else:
        date_formatted = f"Ngày: {result.date}"
        
    new_sheet['U3'].value = date_formatted
    new_sheet['U3'].font = Font(name='Times New Roman', size=11, bold=True)
    new_sheet['U3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 5. Phủ màu theo mức độ ưu tiên để giao diện trực quan và chuyên nghiệp
    priority_colors = {
        1: PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),  # Đỏ nhạt (Silo)
        2: PatternFill(start_color='FFE8CC', end_color='FFE8CC', fill_type='solid'),  # Cam nhạt (Vãng lai)
        3: PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Vàng nhạt (Bù thiếu)
        4: PatternFill(start_color='FFD2D2', end_color='FFD2D2', fill_type='solid'),  # Hồng cảnh báo (Đứt hàng DOH < 3)
        5: PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid'),  # Xanh lá cây nhạt (Forecast thường)
    }
    no_fill = PatternFill(fill_type=None)
    
    # Fonts tiêu chuẩn của nhà máy
    font_regular = Font(name='Times New Roman', size=12, bold=False)
    font_bold = Font(name='Times New Roman', size=12, bold=True)
    
    # 6. Ghi dữ liệu vào lưới 35 dòng (dòng 7 đến dòng 41)
    data_start = 7
    data_end = 41
    
    for r in range(data_start, data_end + 1):
        idx = r - data_start
        item: Optional[DemandItem] = result.items[idx] if idx < len(result.items) else None
        
        # Thiết lập cột STT (cột A)
        cell_stt = new_sheet[f'A{r}']
        cell_stt.value = r - 6
        cell_stt.font = font_bold
        cell_stt.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột TÊN CÁM (cột B)
        cell_name = new_sheet[f'B{r}']
        cell_name.value = item.product_code if item else None
        cell_name.font = font_bold
        cell_name.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột KẾ HOẠCH MẺ (cột C)
        cell_batches = new_sheet[f'C{r}']
        cell_batches.value = item.batches if item else None
        cell_batches.font = font_bold
        cell_batches.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột TỔNG TẤN (cột D) - dùng giá trị thực thay vì công thức
        # vì sản phẩm gộp (SILO + đóng bao) có tons ≠ batches × tpb cố định
        cell_tons = new_sheet[f'D{r}']
        cell_tons.value = item.tons if item else None
        cell_tons.font = font_bold
        cell_tons.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột bao bì (Cột E đến S) và Silo Truck (Cột T)
        cols_pkg = {
            'E': item.higro_25 if item else 0.0,
            'F': item.higro_40 if item else 0.0,
            'G': item.cp_25 if item else 0.0,
            'H': item.cp_40 if item else 0.0,
            'I': item.star_25 if item else 0.0,
            'J': item.star_40 if item else 0.0,
            'K': item.nuvo_25 if item else 0.0,
            'L': item.nuvo_40 if item else 0.0,
            'M': item.bell_25 if item else 0.0,
            'N': item.bell_40 if item else 0.0,
            'O': item.nasa_25 if item else 0.0,
            'P': item.nasa_40 if item else 0.0,
            'Q': item.white_bag_25 if item else 0.0,
            'R': item.white_bag_40 if item else 0.0,
            'S': item.white_bag_50 if item else 0.0,
            'T': item.silo_truck if item else 0.0,
        }
        
        for col_letter, val in cols_pkg.items():
            cell_pkg = new_sheet[f'{col_letter}{r}']
            cell_pkg.value = val if (val and val > 0) else None
            cell_pkg.font = font_regular
            cell_pkg.alignment = Alignment(horizontal='center', vertical='center')
            
        # Thiết lập giá trị trực tiếp cho các cột (tránh lỗi công thức VLOOKUP lệch kiểu dữ liệu hoặc ghi đè ràng buộc thông minh)
        # Cột U: KHÁNG SINH (sử dụng công thức Excel động liên kết với sheet KHÁNG SINH)
        new_sheet[f'U{r}'].value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r}, \'KHÁNG SINH\'!$B$3:$C$2000, 2, 0), "SẠCH (KHÔNG KS)"))'
        new_sheet[f'U{r}'].font = font_regular
        new_sheet[f'U{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột V: LINE CV
        new_sheet[f'V{r}'].value = item.line_cv if (item and item.line_cv) else None
        new_sheet[f'V{r}'].font = font_regular
        new_sheet[f'V{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột W: LINE PK
        new_sheet[f'W{r}'].value = item.line_pk if (item and item.line_pk) else None
        new_sheet[f'W{r}'].font = font_regular
        new_sheet[f'W{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột X: CHÊNH LỆCH
        new_sheet[f'X{r}'].value = f'=IF(D{r}="","",D{r}-SUM(E{r}:T{r}))'
        new_sheet[f'X{r}'].font = font_regular
        new_sheet[f'X{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột Y: Dòng trống phân tách
        new_sheet[f'Y{r}'].value = None
        
        # Cột Z: THỰC HIỆN MẺ
        new_sheet[f'Z{r}'].value = f'=IFERROR(VLOOKUP(B{r},$AC$7:$AD$41,2,0),0)'
        new_sheet[f'Z{r}'].font = font_regular
        new_sheet[f'Z{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột AA: HOÀN THÀNH (%)
        new_sheet[f'AA{r}'].value = f'=IF(C{r}=0,0,Z{r}/C{r}*100)'
        new_sheet[f'AA{r}'].font = font_regular
        new_sheet[f'AA{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột AB: LÝ DO CHI TIẾT
        new_sheet[f'AB{r}'].value = None
        new_sheet[f'AB{r}'].font = font_regular
        new_sheet[f'AB{r}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Cột AC và AD: Dữ liệu thực tế pha trộn (để trống chờ nhập)
        new_sheet[f'AC{r}'].value = None
        new_sheet[f'AC{r}'].font = font_regular
        new_sheet[f'AC{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        new_sheet[f'AD{r}'].value = None
        new_sheet[f'AD{r}'].font = font_regular
        new_sheet[f'AD{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Áp dụng màu nền hàng theo độ ưu tiên
        row_fill = priority_colors.get(item.priority.value) if item else no_fill
        for c in range(1, 31):  # Cột A đến AD (cột 1 đến 30)
            new_sheet.cell(row=r, column=c).fill = row_fill
            
    # Điền công thức VLOOKUP cột U cho các dòng từ 42 đến 450 (bỏ qua các dòng chữ ký và footer từ 42 đến 50)
    for r in range(42, 451):
        if r in {42, 43, 44, 45, 46, 47, 48, 49, 50}:
            continue
        new_sheet[f'U{r}'].value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r}, \'KHÁNG SINH\'!$B$3:$C$2000, 2, 0), "SẠCH (KHÔNG KS)"))'
        new_sheet[f'U{r}'].font = font_regular
        new_sheet[f'U{r}'].alignment = Alignment(horizontal='center', vertical='center')
            
    # 7. Cập nhật các dòng tổng kết và ký duyệt ở chân trang (Footer)
    # Dòng 42: TỔNG (TOTAL)
    new_sheet['A42'].value = 'TỔNG (TOTAL)'
    new_sheet['A42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['A42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['C42'].value = '=IF(SUM(C7:C41)=0,"",SUM(C7:C41))'
    new_sheet['C42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['C42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['D42'].value = '=IF(SUM(D7:D41)=0,"",SUM(D7:D41))'
    new_sheet['D42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['D42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['Z42'].value = '=SUM(Z7:Z41)'
    new_sheet['Z42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['Z42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AA42'].value = '=IF(C42=0,0,(Z42/C42)*100)'
    new_sheet['AA42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AA42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AD42'].value = '=SUM(AD7:AD41)'
    new_sheet['AD42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AD42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AE42'].value = '=$AD$42-$C$42'
    new_sheet['AE42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AE42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AF42'].value = 'SỐ MẺ'
    new_sheet['AF42'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['AF42'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Dòng 43: Sổ sách chất lượng và tỷ lệ blender
    new_sheet['A43'].value = 'QT-SX-01/BM04\nLần ban hành: 04\nNgày hiệu lực: 01/10//2025'
    new_sheet['A43'].font = Font(name='Times New Roman', size=12, bold=False)
    new_sheet['A43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    new_sheet['H43'].value = 'NGƯỜI LẬP KẾ HOẠCH '
    new_sheet['H43'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['H43'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['U43'].value = 'NGƯỜI THẨM TRA'
    new_sheet['U43'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['U43'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AE43'].value = '=IF(C42=0,0,($AD$42/$C$42)*100)'
    new_sheet['AE43'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AE43'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AF43'].value = '%'
    new_sheet['AF43'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['AF43'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Dòng 47: Tên người ký lập kế hoạch
    new_sheet['H47'].value = 'Hồ Đăng Xuân Thành'
    new_sheet['H47'].font = Font(name='Times New Roman', size=9, bold=True)
    new_sheet['H47'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 7.5. Chèn động các mã cám mới và kháng sinh vào sheet 'KHÁNG SINH'
    if 'KHÁNG SINH' in wb.sheetnames:
        ks_sheet = wb['KHÁNG SINH']
        existing_codes = set()
        
        # Tìm dòng cuối cùng thực tế có dữ liệu để tránh ghi đè dữ liệu cũ
        max_filled_row = 2
        for r in range(3, 2000):
            cell = ks_sheet.cell(row=r, column=2)
            val = cell.value
            if val is not None and str(val).strip() != '':
                # Đồng bộ kiểu dữ liệu thành String để tránh lỗi VLOOKUP lệch kiểu (str vs int) trong Excel
                cell.value = str(val).strip()
                max_filled_row = r
                existing_codes.add(str(val).strip().upper().replace(' ', ''))
        empty_row = max_filled_row + 1
                    
        for item in result.items:
            if not item or not item.product_code:
                continue
            p_code = str(item.product_code).strip().upper().replace(' ', '')
            if p_code and p_code not in existing_codes:
                ks_sheet.cell(row=empty_row, column=1).value = empty_row - 2
                ks_sheet.cell(row=empty_row, column=2).value = str(item.product_code).strip()
                ks_sheet.cell(row=empty_row, column=3).value = item.ks_code if item.ks_code else 'KS/ HC (1)/(2)'
                existing_codes.add(p_code)
                empty_row += 1

    # Đồng bộ hóa kiểu dữ liệu cho sheet FEEDCODE để tránh lỗi VLOOKUP lệch kiểu tương tự
    if 'FEEDCODE' in wb.sheetnames:
        fd_sheet = wb['FEEDCODE']
        for r in range(2, fd_sheet.max_row + 1):
            cell = fd_sheet.cell(row=r, column=2)  # Cột B = Tên cám
            val = cell.value
            if val is not None and str(val).strip() != '':
                cell.value = str(val).strip()

    # 8. Xóa toàn bộ các sheet không liên quan để file tinh gọn
    # Chỉ giữ lại sheet kế hoạch ngày mới, FEEDCODE, KHÁNG SINH và LINE PK VÀ CV
    sheets_to_keep = {result.date, 'FEEDCODE', 'KHÁNG SINH', 'LINE PK VÀ CV'}
    for name in list(wb.sheetnames):
        if name not in sheets_to_keep:
            wb.remove(wb[name])
            
    # Sửa lỗi openpyxl copy/delete sheet làm hỏng danh sách fill toàn cục
    from openpyxl.styles.fills import Fill
    cleaned_fills = []
    for f in list(wb._fills):
        if isinstance(f, Fill):
            cleaned_fills.append(f)
    wb._fills = cleaned_fills
            
    # 9. Lưu workbook mới
    os.makedirs(output_dir, exist_ok=True)
    try:
        wb.save(filepath)
        wb.close()
    except PermissionError:
        # File đang mở ở Excel, lưu bản sao mới
        import time
        suffix = int(time.time()) % 1000
        base, ext = os.path.splitext(filepath)
        filepath = f"{base}_v{suffix}{ext}"
        print(f"\n⚠️ CẢNH BÁO: File gốc đang mở trong Excel hoặc bị khóa! Tự động lưu bản sao...")
        wb.save(filepath)
        wb.close()
            
    print(f"\n  💾 Đã xuất KHSX chất lượng cao: {filepath}")
    return filepath


def print_khsx_console(result):
    """In KHSX ra console dạng bảng để kiểm tra nhanh"""
    print(f"\n{'═'*90}")
    print(f"  KẾ HOẠCH SẢN XUẤT - NGÀY {result.date}")
    print(f"{'═'*90}")
    
    print(f"{'STT':>4} {'TÊN CÁM':<12} {'MẺ':>4} {'TẤN':>8} {'BAO BÌ':<30} {'LINE':>6} {'NGUỒN':<10}")
    print(f"{'─'*90}")
    
    for i, item in enumerate(result.items, 1):
        pkg_parts = []
        if item.silo_truck: pkg_parts.append(f"SILO:{item.silo_truck:.1f}")
        if item.white_bag_50: pkg_parts.append(f"WH50:{item.white_bag_50:.1f}")
        if item.higro_25: pkg_parts.append(f"HG:{item.higro_25:.1f}")
        if item.cp_25: pkg_parts.append(f"CP:{item.cp_25:.1f}")
        if item.star_25: pkg_parts.append(f"ST:{item.star_25:.1f}")
        if item.white_bag_25: pkg_parts.append(f"WH25:{item.white_bag_25:.1f}")
        pkg_str = ' '.join(pkg_parts) if pkg_parts else '-'
        
        priority_icon = {1: '🔴', 2: '🟠', 3: '🟡', 4: '🟢'}.get(item.priority.value, '⚪')
        
        print(f"{i:>4} {item.product_code:<12} {item.batches:>4} {item.tons:>8.1f} "
              f"{pkg_str:<30} {item.line_cv:>6} {priority_icon} {item.source:<10}")
    
    print(f"{'─'*90}")
    print(f"     {'TỔNG':<12} {result.total_batches:>4} {result.total_tons:>8.1f}")
    print(f"{'═'*90}")
    
    if result.warnings:
        print(f"\n⚠️ CẢNH BÁO ({len(result.warnings)}):")
        for w in result.warnings:
            print(f"  {w}")
