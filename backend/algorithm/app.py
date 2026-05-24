"""
app.py - Backend Flask server for KHSX Automator
C.P. Vietnam - Chi nhánh Bình Dương
"""

import sys
import os
import io
import datetime
import subprocess
import tempfile
import json
import traceback
from flask import Flask, render_template, jsonify, request, send_from_directory, Response

# Fix encoding cho Windows console
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Thêm thư mục hiện tại vào path
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, CURRENT_DIR)

import config
import data_loader
from models import Priority, PackingType

app = Flask(__name__, template_folder=os.path.join(CURRENT_DIR, 'templates'), static_folder=os.path.join(CURRENT_DIR, 'static'))
app.config['SECRET_KEY'] = 'cp_vietnam_khsx_secret_key'

# Khởi tạo thư mục upload tạm thời
TEMP_UPLOAD_DIR = os.path.join(CURRENT_DIR, 'temp_uploads')
os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def get_file_info(directory, pattern, exact_path=None):
    """Lấy thông tin của file mới nhất theo pattern hoặc đường dẫn chính xác"""
    if exact_path:
        path = exact_path
    else:
        path = data_loader._find_latest_file(directory, pattern)
        
    if not path or not os.path.isfile(path):
        return {
            'exists': False,
            'filename': 'Không tìm thấy',
            'last_modified': '-',
            'size': '-',
            'path': ''
        }
        
    stat = os.stat(path)
    mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
    size_kb = round(stat.st_size / 1024, 1)
    
    return {
        'exists': True,
        'filename': os.path.basename(path),
        'last_modified': mtime.strftime('%d-%m-%Y %H:%M:%S'),
        'size': f"{size_kb} KB",
        'path': path
    }


def _safe_float_val(v):
    if v is None or v == '':
        return ''
    try:
        return float(v)
    except:
        return v


def _safe_float_num(v, default=0.0):
    if v is None or v == '':
        return default
    try:
        return float(v)
    except:
        return default


def _safe_int_val(v):
    if v is None or v == '':
        return ''
    try:
        return int(float(v))
    except:
        return v


# ============================================================
# VIEWS & TEMPLATE ROUTES
# ============================================================

@app.route('/')
def index():
    """Trang chủ Single Page App"""
    return render_template('index.html')


# ============================================================
# API: TRẠNG THÁI NGUỒN DỮ LIỆU
# ============================================================

@app.route('/api/data-status', methods=['GET'])
def get_data_status():
    """Lấy thông tin và trạng thái của 10 nguồn dữ liệu đầu vào"""
    try:
        status = {
            'forecast': get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx'),
            'silo_plan': get_file_info(config.SILO_DIR, '*SILO*.xlsx'),
            'bacang': get_file_info(config.BACANG_DIR, '*CANG*.xlsx'),
            'ffstock': get_file_info(config.FSTOCK_DIR, '*FFSTOCK*.xls*'),
            'tonbon': get_file_info(config.TONBON_DIR, '*ton bon*.*'),
            'empty_bag': get_file_info(config.FSTOCK_DIR, '*EMPTY BAG*.xls*'),
            'congsuat': get_file_info(None, None, exact_path=config.PLAN_FILE),
            'feedcode': get_file_info(None, None, exact_path=config.KHSX_FILE),
            'khangsinh': get_file_info(None, None, exact_path=config.KHSX_FILE),
            'yesterday_plan': get_file_info(None, None, exact_path=config.KHSX_FILE),
            'adjustments': get_file_info(None, None, exact_path=config.QUICK_ADJUST_FILE)
        }
        return jsonify({'success': True, 'status': status})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/latest-target-date', methods=['GET'])
def get_latest_target_date():
    """Tự động quét tìm ngày dữ liệu FFSTOCK mới nhất và tính toán ngày mục tiêu tiếp theo"""
    try:
        import glob
        import re
        
        # 1. Tìm tất cả các file FFSTOCK trong thư mục FSTOCK_DIR
        pattern = os.path.join(config.FSTOCK_DIR, "*FFSTOCK*.xls*")
        files = glob.glob(pattern)
        
        # Lọc bỏ file tạm thời
        files = [f for f in files if not os.path.basename(f).startswith('~$')]
        
        if not files:
            return jsonify({
                'success': False, 
                'message': 'Không tìm thấy bất kỳ file FFSTOCK nào trong thư mục'
            })
            
        # Regex linh hoạt hỗ trợ khoảng trắng
        date_pattern = re.compile(r'FFSTOCK\s*(\d{1,2})\s*[-/:\s]\s*(\d{1,2})\s*[-/:\s]\s*(\d{4})', re.IGNORECASE)
        
        dates = []
        for filepath in files:
            filename = os.path.basename(filepath)
            match = date_pattern.search(filename)
            if match:
                d = int(match.group(1))
                m = int(match.group(2))
                y = int(match.group(3))
                try:
                    dt = datetime.date(y, m, d)
                    dates.append((dt, filepath))
                except Exception:
                    pass
                    
        if not dates:
            return jsonify({
                'success': False,
                'message': 'Không thể trích xuất ngày từ tên các file FFSTOCK'
            })
            
        # Tìm ngày lớn nhất (ngày mới nhất)
        dates.sort(key=lambda x: x[0], reverse=True)
        latest_data_date, latest_file_path = dates[0]
        
        # Tính toán ngày mục tiêu (ngày sau đó 1 ngày)
        target_date = latest_data_date + datetime.timedelta(days=1)
        
        # Format các chuỗi ngày
        data_date_str = latest_data_date.strftime('%d-%m-%Y')
        target_date_str = target_date.strftime('%d-%m-%Y')
        target_date_iso = target_date.strftime('%Y-%m-%d') # cho input date
        
        # 2. Kiểm tra xem file KHSX cho ngày mục tiêu đã được lập chưa
        out_pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{target_date_str}*.xlsx")
        out_files = glob.glob(out_pattern)
        
        has_existing_plan = False
        existing_plan_filename = ""
        
        if out_files:
            out_files.sort(key=os.path.getmtime, reverse=True)
            has_existing_plan = True
            existing_plan_filename = os.path.basename(out_files[0])
            
        return jsonify({
            'success': True,
            'data_date': data_date_str,
            'target_date': target_date_str,
            'target_date_iso': target_date_iso,
            'has_existing_plan': has_existing_plan,
            'existing_plan_filename': existing_plan_filename
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: CHI TIẾT NGUỒN DỮ LIỆU (READ)
# ============================================================

@app.route('/api/data/<category>', methods=['GET'])
def get_detailed_data(category):
    """Đọc dữ liệu chi tiết của từng nguồn đầu vào để hiển thị bảng khoa học"""
    try:
        info = {}
        rows = []
        
        # Lấy file tương ứng
        if category == 'forecast':
            info = get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx')
            if info['exists']:
                items = data_loader.load_forecast(info['path'])
                for it in items:
                    rows.append({
                        'product_code': it.product_code,
                        'packing_size': it.packing_size,
                        'die_size': it.die_size,
                        'dealer_higro': it.dealer_higro,
                        'dealer_cp': it.dealer_cp,
                        'dealer_star': it.dealer_star,
                        'dealer_nuvo': it.dealer_nuvo,
                        'dealer_nasa': it.dealer_nasa,
                        'dealer_total': it.dealer_total,
                        'farm_swine': it.farm_swine,
                        'farm_integrate': it.farm_integrate,
                        'farm_total': it.farm_total,
                        'grand_total_tons': it.grand_total_tons,
                        'silo_tons': it.silo_tons,
                        'total_with_silo': it.total_with_silo
                    })
                    
        elif category == 'silo_plan':
            info = get_file_info(config.SILO_DIR, '*SILO*.xlsx')
            if info['exists']:
                silo_plan = data_loader.load_silo_plan(info['path'])
                # silo_plan: dict {day → {product → tons}}
                all_products = sorted(list(set(p for d in silo_plan.values() for p in d)))
                for p in all_products:
                    row = {'product_code': p}
                    for day in range(1, 7):
                        row[f'day_{day}'] = round(silo_plan[day].get(p, 0.0), 1)
                    row['total'] = round(sum(row[f'day_{d}'] for d in range(1, 7)), 1)
                    rows.append(row)
                    
        elif category == 'bacang':
            info = get_file_info(config.BACANG_DIR, '*CANG*.xlsx')
            if info['exists']:
                bacang = data_loader.load_bacang(info['path'])
                # bacang: dict {day → {product → tons}}
                all_products = sorted(list(set(p for d in bacang.values() for p in d)))
                for p in all_products:
                    row = {'product_code': p}
                    for day in range(1, 7):
                        row[f'day_{day}'] = round(bacang[day].get(p, 0.0), 1)
                    row['total'] = round(sum(row[f'day_{d}'] for d in range(1, 7)), 1)
                    rows.append(row)
                    
        elif category == 'ffstock':
            info = get_file_info(config.FSTOCK_DIR, '*FFSTOCK*.xls*')
            if info['exists']:
                ffstock = data_loader.load_ffstock(info['path'])
                details = data_loader.load_ffstock_details(info['path'])
                
                # Load forecast cho Plan column
                forecast_map = {}
                forecast_info = get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx')
                if forecast_info['exists']:
                    try:
                        fc_items = data_loader.load_forecast(forecast_info['path'])
                        for it in fc_items:
                            forecast_map[it.product_code] = {
                                'plan': it.total_with_silo if it.total_with_silo else it.grand_total_tons,
                                'packing_size': it.packing_size,
                                'die_size': it.die_size
                            }
                    except:
                        pass
                
                # Load tên cám từ FEEDCODE sheet (nếu có)
                product_names = {}
                try:
                    congsuat = data_loader.load_congsuat(config.PLAN_FILE)
                    for spec in congsuat:
                        if spec.product_code and spec.formular_code:
                            product_names[spec.product_code] = spec.formular_code
                except:
                    pass
                
                # Merge all data
                all_products = sorted(list(set(list(ffstock.keys()) + list(details.keys()))))
                for p in all_products:
                    det = details.get(p, {})
                    stock_tons = round(ffstock.get(p, 0.0), 1)
                    daily_avg = round(det.get('daily_sales_tons', 0.0), 1)
                    doh_val = det.get('doh', None)
                    doh = round(doh_val, 1) if doh_val is not None else None
                    
                    # Plan từ forecast
                    fc = forecast_map.get(p, {})
                    plan = round(fc.get('plan', 0.0), 1) if fc.get('plan') else 0
                    
                    # Day5 = forecast / 6 * 5 (5 ngày tới cần bao nhiêu)
                    day5 = round(plan / 6 * 5, 1) if plan > 0 else 0
                    
                    # Phân loại vật nuôi
                    animal = classify_animal_type(p)
                    
                    # KQ GC2: so sánh Stock với nhu cầu
                    kq_gc2 = ''
                    if daily_avg > 0 and stock_tons > 0:
                        kq_gc2 = round(stock_tons - plan, 1) if plan > 0 else 0
                    
                    # Ghi chú tự động
                    ghi_chu = ''
                    if doh is not None and doh < 1:
                        ghi_chu = f'⚠️ CẦN SX NGAY'
                    elif doh is not None and doh < 3:
                        ghi_chu = f'📋 DOH thấp'
                    
                    rows.append({
                        'product_code': p,
                        'product_name': product_names.get(p, det.get('product_name', '')),
                        'animal_type': animal,
                        'animal_label': VAT_NUOI_LABELS.get(animal, 'HEO'),
                        'animal_color': VAT_NUOI_COLORS.get(animal, '#FF6B6B'),
                        'stock_tons': stock_tons,
                        'safety_stock_tons': round(det.get('safety_stock_tons', 0.0), 1),
                        'daily_sales_tons': daily_avg,
                        'doh': doh,
                        'plan': plan,
                        'day5': day5,
                        'kq_gc2': kq_gc2,
                        'warning': det.get('warning', ''),
                        'ghi_chu': ghi_chu
                    })
                    
        elif category == 'tonbon':
            info = get_file_info(config.TONBON_DIR, '*ton bon*.*')
            if info['exists']:
                tonbon = data_loader.load_tonbon(info['path'])
                # tonbon: dict {product → tons}
                for p, t in sorted(tonbon.items()):
                    rows.append({
                        'product_code': p,
                        'tons': round(t, 1)
                    })
                    
        elif category == 'empty_bag':
            info = get_file_info(config.FSTOCK_DIR, '*EMPTY BAG*.xls*')
            if info['exists']:
                empty_bag = data_loader.load_empty_bag(info['path'])
                # empty_bag: dict {product → {brand → bags}}
                for p, brands in sorted(empty_bag.items()):
                    rows.append({
                        'product_code': p,
                        'HIGRO': brands.get('HIGRO', 0),
                        'CP': brands.get('CP', 0),
                        'STAR': brands.get('STAR', 0),
                        'NASA': brands.get('NASA', 0),
                        'NUVO': brands.get('NUVO', 0),
                        'FARM': brands.get('FARM', 0)
                    })
                    
        elif category == 'congsuat':
            info = get_file_info(None, None, exact_path=config.PLAN_FILE)
            if info['exists']:
                congsuat = data_loader.load_congsuat(info['path'])
                # congsuat: dict {product → ProductSpec}
                for p, spec in sorted(congsuat.items()):
                    rows.append({
                        'product_code': spec.product_code,
                        'formular_code': spec.formular_code,
                        'die_size': spec.die_size,
                        'ton_per_batch': spec.ton_per_batch,
                        'line_cv': spec.line_cv,
                        'line_pk': spec.line_pk,
                        'ks_code': spec.ks_code
                    })
                    
        elif category == 'feedcode':
            info = get_file_info(None, None, exact_path=config.KHSX_FILE)
            if info['exists']:
                feedcode = data_loader.load_feedcode(info['path'])
                # feedcode: dict {product → {line_cv, line_pk}}
                for p, lines in sorted(feedcode.items()):
                    rows.append({
                        'product_code': p,
                        'line_cv': lines.get('line_cv', ''),
                        'line_pk': lines.get('line_pk', '')
                    })
                    
        elif category == 'khangsinh':
            info = get_file_info(None, None, exact_path=config.KHSX_FILE)
            if info['exists']:
                khangsinh = data_loader.load_khangsinh(info['path'])
                # khangsinh: dict {product → ks_code}
                for p, ks in sorted(khangsinh.items()):
                    rows.append({
                        'product_code': p,
                        'ks_code': ks
                    })
                    
        elif category == 'yesterday_plan':
            info = get_file_info(None, None, exact_path=config.KHSX_FILE)
            if info['exists']:
                # Mặc định lấy ngày hôm qua dựa trên ngày hiện tại
                t = datetime.date.today()
                yesterday_day = t.day - 1
                if yesterday_day < 1: yesterday_day = 1
                
                yesterday = data_loader.load_khsx_yesterday(info['path'], yesterday_day)
                # yesterday: dict {product → dict}
                for p, det in sorted(yesterday.items()):
                    rows.append({
                        'product_code': p,
                        'planned_batches': det.get('planned_batches', 0),
                        'actual_batches': det.get('actual_batches', 0),
                        'planned_tons': round(det.get('planned_tons', 0.0), 1),
                        'actual_tons': round(det.get('actual_tons', 0.0), 1),
                        'status': det.get('status', '')
                    })
                    
        elif category == 'adjustments':
            # Trả về dưới dạng 4 danh sách riêng biệt cho các bảng editor
            info = get_file_info(None, None, exact_path=config.QUICK_ADJUST_FILE)
            if info['exists']:
                adj = data_loader.load_quick_adjustments(info['path'])
                
                # Flat additions
                additions = []
                for item in adj.get('additions', []):
                    additions.append({
                        'product_code': item.get('product_code', ''),
                        'tons': item.get('tons', 0.0),
                        'packing_size': item.get('packing_size', ''),
                        'priority': item.get('priority', ''),
                        'force_batches': item.get('force_batches', ''),
                        'force_tpb': item.get('force_tpb', ''),
                        'note': item.get('note', '')
                    })
                    
                # Flat cancellations
                cancellations = []
                for p, cancel_type in adj.get('cancellations', {}).items():
                    cancellations.append({
                        'product_code': p,
                        'cancel_type': cancel_type,
                        'note': ''
                    })
                    
                # Flat substitutions
                substitutions = []
                for old, new in adj.get('substitutions', {}).items():
                    substitutions.append({
                        'old_code': old,
                        'new_code': new,
                        'note': ''
                    })
                    
                # Flat bag substitutions
                bag_substitutions = []
                for p, mapping in adj.get('bag_substitutions', {}).items():
                    for old_bag, new_bag in mapping.items():
                        bag_substitutions.append({
                            'product_code': p,
                            'old_bag': old_bag,
                            'new_bag': new_bag,
                            'note': ''
                        })
                        
                return jsonify({
                    'success': True,
                    'file_info': info,
                    'data': {
                        'additions': additions,
                        'cancellations': cancellations,
                        'substitutions': substitutions,
                        'bag_substitutions': bag_substitutions
                    }
                })
        
        return jsonify({
            'success': True,
            'file_info': info,
            'data': rows
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: UPLOAD FILE ĐẦU VÀO
# ============================================================

@app.route('/api/upload/<category>', methods=['POST'])
def upload_file(category):
    """Admin tải lên báo cáo Excel mới. Tự động lưu vào đúng thư mục cấu hình."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Không tìm thấy file tải lên'})
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Tên file không hợp lệ'})
            
        # Kiểm tra phần mở rộng file Excel
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ['.xlsx', '.xls', '.xlsm']:
            return jsonify({'success': False, 'message': 'Chỉ chấp nhận file Excel (.xlsx, .xls, .xlsm)'})
            
        # Xác định thư mục đích dựa theo config
        target_dir = None
        prefix = f"UPLOAD_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_"
        
        if category == 'forecast':
            target_dir = config.FORECAST_DIR
        elif category == 'silo_plan':
            target_dir = config.SILO_DIR
        elif category == 'bacang':
            target_dir = config.BACANG_DIR
        elif category == 'ffstock':
            target_dir = config.FSTOCK_DIR
        elif category == 'tonbon':
            target_dir = config.TONBON_DIR
        elif category == 'empty_bag':
            target_dir = config.FSTOCK_DIR
        elif category == 'congsuat':
            # file cố định Plan.xlsm
            target_dir = config.PLAN_DIR
            prefix = ""  # Ghi đè trực tiếp
            filename = "Plan.xlsm"
        elif category == 'feedcode' or category == 'khangsinh' or category == 'yesterday_plan':
            # file cố định KHSX THANG 5-20261.xlsm
            target_dir = config.DATA_DIR
            prefix = ""  # Ghi đè trực tiếp
            filename = "KHSX THANG 5-20261.xlsm"
        elif category == 'adjustments':
            target_dir = config.DATA_DIR
            prefix = ""  # Ghi đè trực tiếp
            filename = "DIEU_CHINH_NHANH.xlsx"
            
        if not target_dir:
            return jsonify({'success': False, 'message': f'Danh mục upload không hợp lệ: {category}'})
            
        # Tạo thư mục nếu chưa có
        os.makedirs(target_dir, exist_ok=True)
        
        if prefix != "":
            filename = prefix + file.filename
            
        target_path = os.path.join(target_dir, filename)
        
        # Nếu ghi đè file Plan/KHSX gốc, sao lưu file cũ trước
        if prefix == "" and os.path.isfile(target_path):
            backup_path = target_path + f".bak_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                os.rename(target_path, backup_path)
            except Exception as ex:
                print(f"Không thể sao lưu file cũ: {ex}")
                
        file.save(target_path)
        
        # Trả về thông tin file mới
        info = get_file_info(None, None, exact_path=target_path)
        return jsonify({
            'success': True,
            'message': f"Đã upload thành công và lưu vào {os.path.basename(target_path)}",
            'file_info': info
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: QUẢN TRỊ ĐIỀU CHỈNH NHANH (DIEU_CHINH_NHANH.xlsx)
# ============================================================

@app.route('/api/adjustments/save', methods=['POST'])
def save_adjustments():
    """Lưu trực tiếp dữ liệu từ bảng editor vào DIEU_CHINH_NHANH.xlsx"""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'message': 'Dữ liệu trống'})
            
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        file_path = config.QUICK_ADJUST_FILE
        wb = Workbook()
        
        # Định dạng style mẫu CP
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        # 1. THEM_MOI_HOAC_SUA
        ws1 = wb.active
        ws1.title = "THEM_MOI_HOAC_SUA"
        headers1 = ["MÃ CÁM", "TẤN", "QUY CÁCH", "LOẠI ƯU TIÊN", "ÉP SỐ MẺ", "ÉP TẤN/MẺ", "GHI CHÚ (Tùy chọn)"]
        ws1.append(headers1)
        
        additions = data.get('additions', [])
        for item in additions:
            ws1.append([
                str(item.get('product_code', '')).strip().upper(),
                _safe_float_val(item.get('tons')),
                str(item.get('packing_size', '')).strip().upper(),
                str(item.get('priority', '')).strip().upper(),
                _safe_int_val(item.get('force_batches')),
                _safe_float_val(item.get('force_tpb')),
                item.get('note', '')
            ])
            
        # 2. HUY_KHSX
        ws2 = wb.create_sheet("HUY_KHSX")
        headers2 = ["MÃ CÁM", "LOẠI HỦY", "GHI CHÚ"]
        ws2.append(headers2)
        cancellations = data.get('cancellations', [])
        for item in cancellations:
            ws2.append([
                str(item.get('product_code', '')).strip().upper(),
                str(item.get('cancel_type', '')).strip().upper(),
                item.get('note', '')
            ])
            
        # 3. THAY_THE_MA_CAM
        ws3 = wb.create_sheet("THAY_THE_MA_CAM")
        headers3 = ["MÃ CŨ", "MÃ MỚI", "GHI CHÚ"]
        ws3.append(headers3)
        substitutions = data.get('substitutions', [])
        for item in substitutions:
            ws3.append([
                str(item.get('old_code', '')).strip().upper(),
                str(item.get('new_code', '')).strip().upper(),
                item.get('note', '')
            ])
            
        # 4. THAY_THE_BAO_BI
        ws4 = wb.create_sheet("THAY_THE_BAO_BI")
        headers4 = ["MÃ CÁM", "BAO GỐC", "BAO THAY THẾ", "GHI CHÚ"]
        ws4.append(headers4)
        bag_substitutions = data.get('bag_substitutions', [])
        for item in bag_substitutions:
            ws4.append([
                str(item.get('product_code', '')).strip().upper(),
                str(item.get('old_bag', '')).strip().upper(),
                str(item.get('new_bag', '')).strip().upper(),
                item.get('note', '')
            ])
            
        # Apply header styling cho tất cả sheets
        for sheet in wb.worksheets:
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                
        # Sao lưu file cũ trước khi lưu
        if os.path.isfile(file_path):
            backup_path = file_path + f".bak_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try: os.rename(file_path, backup_path)
            except: pass
            
        wb.save(file_path)
        
        info = get_file_info(None, None, exact_path=file_path)
        return jsonify({
            'success': True,
            'message': 'Đã ghi nhận thay đổi và cập nhật DIEU_CHINH_NHANH.xlsx thành công!',
            'file_info': info
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: CHẠY THUẬT TOÁN KHSX (LIVE SSE STREAM)
# ============================================================

@app.route('/api/generate-plan', methods=['POST'])
def generate_plan():
    """Nhận ngày tính toán, chạy script tối ưu hóa KHSX dưới dạng luồng Stream logs SSE"""
    try:
        req_data = request.json or {}
        target_date_str = req_data.get('date')  # YYYY-MM-DD
        walkin_orders = req_data.get('walkin_orders', [])
        
        if not target_date_str:
            return jsonify({'success': False, 'message': 'Ngày không hợp lệ'})
            
        # Parse ngày sang định dạng YYYYMMDD cho script
        dt = datetime.datetime.strptime(target_date_str, '%Y-%m-%d')
        date_param = dt.strftime('%Y%m%d')
        date_dmy = dt.strftime('%d-%m-%Y')
        
        # 1. Tạo file đơn vãng lai tạm thời nếu có
        walkin_file_path = None
        if walkin_orders:
            temp_fd, walkin_file_path = tempfile.mkstemp(suffix='.csv', text=True)
            with open(walkin_file_path, 'w', encoding='utf-8', newline='') as csvfile:
                writer = csv.writer(csvfile)
                for order in walkin_orders:
                    writer.writerow([
                        str(order.get('product', '')).strip().upper(),
                        float(order.get('tons', 0.0)),
                        str(order.get('packing_size', '25')).strip()
                    ])
            os.close(temp_fd)
            
        # 2. Định nghĩa hàm generator để stream dữ liệu log console
        def sse_generator():
            # Chạy script Python khsx_auto.py dưới dạng tiến trình con
            cmd = ['py', 'khsx_auto.py', '--date', date_param]
            if walkin_file_path:
                cmd.extend(['--walkin', walkin_file_path])
                
            yield "data: " + json.dumps({'type': 'log', 'text': f"🚀 Khởi động chương trình KHSX tự động cho ngày {date_dmy}..."}) + "\n\n"
            
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                cwd=CURRENT_DIR
            )
            
            # Đọc output từng dòng
            for line in process.stdout:
                line_clean = line.rstrip()
                if line_clean:
                    # Gửi log từng dòng về frontend
                    yield "data: " + json.dumps({'type': 'log', 'text': line_clean}) + "\n\n"
                    
            process.wait()
            
            # Xóa file đơn vãng lai tạm thời
            if walkin_file_path and os.path.exists(walkin_file_path):
                try: os.remove(walkin_file_path)
                except: pass
                
            if process.returncode == 0:
                # Tìm file KHSX đầu ra vừa tạo
                import glob
                pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{date_dmy}*.xlsx")
                files = glob.glob(pattern)
                
                output_filename = ""
                if files:
                    files.sort(key=os.path.getmtime, reverse=True)
                    output_filename = os.path.basename(files[0])
                    
                yield "data: " + json.dumps({
                    'type': 'complete',
                    'success': True,
                    'output_file': output_filename,
                    'message': f"🎉 Đã lập KHSX thành công cho ngày {date_dmy}!"
                }) + "\n\n"
            else:
                yield "data: " + json.dumps({
                    'type': 'complete',
                    'success': False,
                    'message': f"❌ Lỗi thực thi thuật toán (Mã thoát: {process.returncode}). Vui lòng kiểm tra lại dữ liệu Excel đầu vào!"
                }) + "\n\n"
                
        return Response(sse_generator(), mimetype='text/event-stream')
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# HELPER: PHÂN LOẠI VẬT NUÔI TỰ ĐỘNG TỪ MÃ CÁM
# ============================================================

# Bảng mapping vật nuôi
VAT_NUOI_LABELS = {'H': 'HEO', 'G': 'GÀ', 'B': 'BÒ', 'V': 'VỊT', 'C': 'CÚT', 'D': 'DÊ'}
VAT_NUOI_COLORS = {'H': '#FF6B6B', 'G': '#4ECDC4', 'B': '#45B7D1', 'V': '#96CEB4', 'C': '#FFEAA7', 'D': '#DDA0DD'}

def classify_animal_type(product_code: str) -> str:
    """
    Phân loại vật nuôi tự động dựa theo prefix mã cám.
    Quy tắc: 1xx-5xx = Heo, 6xx-7xx = Gà, 8xx = Vịt, 9xx = Bò/Dê/Cút
    Trả về mã 1 ký tự: H/G/B/V/C/D
    """
    code = str(product_code).strip().upper()
    if not code:
        return 'H'
    
    # Lấy ký tự đầu tiên (số)
    first_char = code[0]
    
    if first_char in ('1', '2', '3', '4', '5'):
        return 'H'  # HEO
    elif first_char in ('6', '7'):
        return 'G'  # GÀ
    elif first_char == '8':
        # 8xx có thể là Vịt hoặc đặc biệt
        if code.startswith('85') or code.startswith('86') or code.startswith('87'):
            return 'V'  # VỊT
        return 'V'  # Mặc định 8xx = Vịt
    elif first_char == '9':
        # 9xx: Bò, Cút, Dê - cần kiểm tra chi tiết hơn
        if code.startswith('92') or code.startswith('93'):
            return 'B'  # BÒ
        elif code.startswith('95'):
            return 'C'  # CÚT
        elif code.startswith('96') or code.startswith('97'):
            return 'D'  # DÊ
        return 'B'  # Mặc định 9xx = Bò
    else:
        # Mã không bắt đầu bằng số: kiểm tra keyword
        if 'DUCK' in code or 'VIT' in code:
            return 'V'
        elif 'QUAIL' in code or 'CUT' in code:
            return 'C'
        elif 'GOAT' in code or 'DE' in code:
            return 'D'
        elif 'CATTLE' in code or 'BO' in code or 'DAIRY' in code:
            return 'B'
        elif 'BROILER' in code or 'LAYER' in code or 'GA' in code or 'CHICK' in code:
            return 'G'
        return 'H'  # Mặc định = Heo (chiếm đa số)


def load_doh_data_for_sequence(sequence_items):
    """
    Tải dữ liệu DOH (Days on Hand) cho các sản phẩm trong mixer sequence.
    Nguồn 1: FFSTOCK details (cột DOH có sẵn)
    Nguồn 2: Tính từ forecast = stock / (forecast_week / 6)
    Trả về dict {product_code → {doh, stock, daily_avg}}
    """
    doh_map = {}
    
    try:
        # Nguồn 1: Đọc DOH từ FFSTOCK 
        ffstock_info = get_file_info(config.FSTOCK_DIR, '*FFSTOCK*.xls*')
        if ffstock_info['exists']:
            ffstock = data_loader.load_ffstock(ffstock_info['path'])
            details = data_loader.load_ffstock_details(ffstock_info['path'])
            
            for p in set(list(ffstock.keys()) + list(details.keys())):
                det = details.get(p, {})
                stock_tons = ffstock.get(p, 0.0)
                doh_val = det.get('doh', None)
                daily_avg = det.get('daily_sales_tons', 0.0)
                
                doh_map[p] = {
                    'stock': round(stock_tons, 1),
                    'doh': round(doh_val, 1) if doh_val and doh_val > 0 else None,
                    'daily_avg': round(daily_avg, 1) if daily_avg else 0.0
                }
        
        # Nguồn 2: Bổ sung DOH từ forecast cho sản phẩm chưa có
        forecast_info = get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx')
        if forecast_info['exists']:
            items = data_loader.load_forecast(forecast_info['path'])
            for it in items:
                pc = it.product_code
                forecast_week = it.total_with_silo if it.total_with_silo else it.grand_total_tons
                
                if pc in doh_map and doh_map[pc]['doh'] is None and forecast_week and forecast_week > 0:
                    stock = doh_map[pc].get('stock', 0.0)
                    daily_forecast = forecast_week / 6  # 6 ngày làm việc
                    if daily_forecast > 0:
                        doh_map[pc]['doh'] = round(stock / daily_forecast, 1)
                        doh_map[pc]['daily_avg'] = round(daily_forecast, 1)
                elif pc not in doh_map and forecast_week and forecast_week > 0:
                    daily_forecast = forecast_week / 6
                    doh_map[pc] = {
                        'stock': 0.0,
                        'doh': 0.0,
                        'daily_avg': round(daily_forecast, 1)
                    }
                    
    except Exception as e:
        print(f"⚠️ Lỗi khi tải DOH data: {e}")
    
    return doh_map


# ============================================================
# API: TẢI CHI TIẾT KẾ HOẠCH ĐÃ LẬP (PL, PACKAGING, SEQUENCE)
# ============================================================

@app.route('/api/plan-details/<date_str>', methods=['GET'])
def get_plan_details(date_str):
    """
    Đọc ngược dữ liệu từ file Excel kết quả KHSX vừa tạo 
    để hiển thị trực tiếp lên UI (Kế hoạch Pellet Line, Kế hoạch Đóng bao, Mixer Sequence)
    """
    try:
        # date_str format: dd-mm-yyyy (VD: 21-05-2026)
        import glob
        pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{date_str}*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            return jsonify({'success': False, 'message': f'Không tìm thấy file kế hoạch sản xuất cho ngày {date_str}'})
            
        files.sort(key=os.path.getmtime, reverse=True)
        file_path = files[0]
        
        # Đọc dữ liệu từ file Excel KHSX kết quả bằng openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active # Sheet chính
        
        # Xây dựng danh sách mapping line máy mặc định từ sheet FEEDCODE làm fallback
        feedcode_mapping = {}
        if 'FEEDCODE' in wb.sheetnames:
            fc_sheet = wb['FEEDCODE']
            for fcr in range(2, 500):
                f_name = fc_sheet.cell(row=fcr, column=2).value
                f_cv = fc_sheet.cell(row=fcr, column=3).value
                f_pk = fc_sheet.cell(row=fcr, column=4).value
                if f_name:
                    key = str(f_name).replace(" ", "").strip().upper()
                    feedcode_mapping[key] = {
                        'line_cv': str(f_cv or '').strip(),
                        'line_pk': str(f_pk or '').strip()
                    }
        
        # Xây dựng danh sách mapping kháng sinh tự động đề phòng công thức Excel chưa chạy
        ks_mapping = {}
        if 'KHÁNG SINH' in wb.sheetnames:
            ks_sheet = wb['KHÁNG SINH']
            for ksr in range(3, 2000):
                k_prod = ks_sheet.cell(row=ksr, column=2).value
                k_code = ks_sheet.cell(row=ksr, column=3).value
                if k_prod:
                    ks_mapping[str(k_prod).strip().upper()] = str(k_code).strip()
                    
        # 1. Đọc Mixer Sequence (Bảng kế hoạch sản xuất tổng hợp)
        # Bắt đầu từ dòng 7 đến 41
        sequence = []
        summary = {
            'total_batches': 0,
            'total_tons': 0.0,
            'product_count': 0,
            'warnings': []
        }
        
        for r in range(7, 42): # Dòng 7 đến 41
            prod = ws.cell(row=r, column=2).value # B
            if not prod:
                continue
                
            prod_code = str(prod).strip().upper()
            if prod_code in {'TỔNG CỘNG', 'TOTAL', ''}:
                continue
                
            batches = _safe_int_val(ws.cell(row=r, column=3).value)
            if not batches or batches == '':
                batches = 0
                
            # Đọc các cột bao bì để tự tính tons và quy cách an toàn
            higro_25 = round(_safe_float_num(ws.cell(row=r, column=5).value), 1)
            higro_40 = round(_safe_float_num(ws.cell(row=r, column=6).value), 1)
            cp_25 = round(_safe_float_num(ws.cell(row=r, column=7).value), 1)
            cp_40 = round(_safe_float_num(ws.cell(row=r, column=8).value), 1)
            star_25 = round(_safe_float_num(ws.cell(row=r, column=9).value), 1)
            star_40 = round(_safe_float_num(ws.cell(row=r, column=10).value), 1)
            nuvo_25 = round(_safe_float_num(ws.cell(row=r, column=11).value), 1)
            nuvo_40 = round(_safe_float_num(ws.cell(row=r, column=12).value), 1)
            bell_25 = round(_safe_float_num(ws.cell(row=r, column=13).value), 1)
            bell_40 = round(_safe_float_num(ws.cell(row=r, column=14).value), 1)
            nasa_25 = round(_safe_float_num(ws.cell(row=r, column=15).value), 1)
            nasa_40 = round(_safe_float_num(ws.cell(row=r, column=16).value), 1)
            white_25 = round(_safe_float_num(ws.cell(row=r, column=17).value), 1)
            white_40 = round(_safe_float_num(ws.cell(row=r, column=18).value), 1)
            white_50 = round(_safe_float_num(ws.cell(row=r, column=19).value), 1)
            silo_truck = round(_safe_float_num(ws.cell(row=r, column=20).value), 1)
            
            # Tính toán tons từ công thức nếu openpyxl trả về None/rỗng do chưa mở file
            raw_tons = ws.cell(row=r, column=4).value
            if raw_tons is None or raw_tons == '' or (isinstance(raw_tons, str) and raw_tons.startswith('=')):
                # Tự tính dựa trên batches và mã cám giống công thức Excel
                if batches > 0:
                    if prod_code.startswith('550') or prod_code.startswith('551') or prod_code == '325F':
                        tons = float(batches) * 8.0
                    else:
                        tons = float(batches) * 8.4
                else:
                    tons = 0.0
            else:
                tons = _safe_float_num(raw_tons)
            
            tons = round(tons, 1)

            # Xác định quy cách đóng gói (packing_size) - Hỗ trợ quy cách ghép Silo + Bao
            has_silo = (silo_truck > 0)
            has_bag_25 = (higro_25 > 0 or cp_25 > 0 or star_25 > 0 or nuvo_25 > 0 or bell_25 > 0 or nasa_25 > 0 or white_25 > 0)
            has_bag_40 = (higro_40 > 0 or cp_40 > 0 or star_40 > 0 or nuvo_40 > 0 or bell_40 > 0 or nasa_40 > 0 or white_40 > 0)
            has_bag_50 = (white_50 > 0)
            
            if has_silo and (has_bag_25 or has_bag_40 or has_bag_50):
                bag_types = []
                if has_bag_25: bag_types.append("Bao 25")
                if has_bag_40: bag_types.append("Bao 40")
                if has_bag_50: bag_types.append("Bao 50")
                packing_size = "Silo + " + " + ".join(bag_types)
            elif has_silo:
                packing_size = 'M'
            elif has_bag_50:
                packing_size = '50'
            elif has_bag_40:
                packing_size = '40'
            else:
                packing_size = '25'
                
            # Phân rã tấn theo quy cách để truyền lên UI
            silo_tons = round(silo_truck, 1)
            bag_25_tons = round(higro_25 + cp_25 + star_25 + nuvo_25 + bell_25 + nasa_25 + white_25, 1)
            bag_40_tons = round(higro_40 + cp_40 + star_40 + nuvo_40 + bell_40 + nasa_40 + white_40, 1)
            bag_50_tons = round(white_50, 1)
                
            line_cv = str(ws.cell(row=r, column=22).value or '').strip()
            line_pk = str(ws.cell(row=r, column=23).value or '').strip()
            
            # Khôi phục line_cv từ feedcode_mapping nếu trống
            if not line_cv or line_cv == 'None':
                line_cv = feedcode_mapping.get(prod_code, {}).get('line_cv', '')
            if line_cv and line_cv != 'None':
                if line_cv.isdigit():
                    line_cv = f"PL{line_cv}"
                elif not line_cv.upper().startswith('PL') and not line_cv.upper() == 'MASH':
                    line_cv = f"PL{line_cv}"
                    
            # Khôi phục line_pk từ feedcode_mapping nếu trống
            if not line_pk or line_pk == 'None':
                line_pk = feedcode_mapping.get(prod_code, {}).get('line_pk', '')
            if has_silo and (not line_pk or line_pk == 'None'):
                line_pk = 'SILO'
            
            # Đọc kháng sinh (cột 21 - U), nếu là công thức chưa chạy thì tự VLOOKUP bằng Python
            raw_ks = ws.cell(row=r, column=21).value
            if raw_ks is None or raw_ks == '' or (isinstance(raw_ks, str) and raw_ks.startswith('=')):
                ks_code = ks_mapping.get(prod_code, 'SẠCH (KHÔNG KS)')
            else:
                ks_code = str(raw_ks).strip()
            
            # Cấp độ kháng sinh để phối màu
            ks_level = 1
            if ks_code and ks_code != 'SẠCH (KHÔNG KS)' and ks_code != 'Sạch không KS':
                ks_level = 5 # Đặt mức tượng trưng
                
            sequence.append({
                'product_code': prod_code,
                'batches': batches,
                'tons': tons,
                'packing_size': packing_size,
                'silo_tons': silo_tons,
                'bag_25_tons': bag_25_tons,
                'bag_40_tons': bag_40_tons,
                'bag_50_tons': bag_50_tons,
                'higro_25': higro_25,
                'higro_40': higro_40,
                'cp_25': cp_25,
                'cp_40': cp_40,
                'star_25': star_25,
                'star_40': star_40,
                'nuvo_25': nuvo_25,
                'nuvo_40': nuvo_40,
                'bell_25': bell_25,
                'bell_40': bell_40,
                'nasa_25': nasa_25,
                'nasa_40': nasa_40,
                'white_25': white_25,
                'white_40': white_40,
                'white_50': white_50,
                'silo_truck': silo_truck,
                'line_cv': line_cv,
                'line_pk': line_pk,
                'ks_code': ks_code,
                'ks_level': ks_level
            })
            
            summary['total_batches'] += batches
            summary['total_tons'] += tons
            summary['product_count'] += 1
            
        summary['total_tons'] = round(summary['total_tons'], 1)
        
        # Enrichment: Gắn DOH + phân loại vật nuôi cho mỗi sản phẩm
        doh_map = load_doh_data_for_sequence(sequence)
        
        for item in sequence:
            pc = item['product_code']
            
            # Gắn DOH
            doh_info = doh_map.get(pc, {})
            item['doh'] = doh_info.get('doh', None)
            item['stock'] = doh_info.get('stock', 0.0)
            item['daily_avg'] = doh_info.get('daily_avg', 0.0)
            

        
        # Đọc cảnh báo từ dòng 46 trở đi

        for r in range(46, 60):
            cell_val = ws.cell(row=r, column=2).value
            if cell_val and str(cell_val).startswith('⚠️'):
                summary['warnings'].append(str(cell_val).strip())
                
        # 2. Xây dựng Kế hoạch Pellet Line (Phân chia theo PL1 -> PL7 và Mash)
        # Lọc sequence theo line_cv
        pl_plans = {f'PL{i}': [] for i in range(1, 8)}
        pl_plans['MASH'] = []
        
        for item in sequence:
            line = item['line_cv'].upper().replace(' ', '')
            if 'PL1' in line: pl_plans['PL1'].append(item)
            elif 'PL2' in line: pl_plans['PL2'].append(item)
            elif 'PL3' in line: pl_plans['PL3'].append(item)
            elif 'PL4' in line: pl_plans['PL4'].append(item)
            elif 'PL5' in line: pl_plans['PL5'].append(item)
            elif 'PL6' in line: pl_plans['PL6'].append(item)
            elif 'PL7' in line: pl_plans['PL7'].append(item)
            elif 'MASH' in line or item['packing_size'] == 'M': pl_plans['MASH'].append(item)
            
        # 3. Kế hoạch Đóng bao (Packaging Matrix)
        packaging_list = []
        for r in range(7, 42): # Dòng 7 đến 41
            prod = ws.cell(row=r, column=2).value # B
            if not prod: continue
            prod_code = str(prod).strip().upper()
            if prod_code in {'TỔNG CỘNG', 'TOTAL', ''}: continue
            
            # Đọc các cột bao bì và silo truck bằng _safe_float_num
            higro_25 = _safe_float_num(ws.cell(row=r, column=5).value)
            higro_40 = _safe_float_num(ws.cell(row=r, column=6).value)
            cp_25 = _safe_float_num(ws.cell(row=r, column=7).value)
            cp_40 = _safe_float_num(ws.cell(row=r, column=8).value)
            star_25 = _safe_float_num(ws.cell(row=r, column=9).value)
            star_40 = _safe_float_num(ws.cell(row=r, column=10).value)
            nuvo_25 = _safe_float_num(ws.cell(row=r, column=11).value)
            nuvo_40 = _safe_float_num(ws.cell(row=r, column=12).value)
            bell_25 = _safe_float_num(ws.cell(row=r, column=13).value)
            bell_40 = _safe_float_num(ws.cell(row=r, column=14).value)
            nasa_25 = _safe_float_num(ws.cell(row=r, column=15).value)
            nasa_40 = _safe_float_num(ws.cell(row=r, column=16).value)
            white_25 = _safe_float_num(ws.cell(row=r, column=17).value)
            white_40 = _safe_float_num(ws.cell(row=r, column=18).value)
            white_50 = _safe_float_num(ws.cell(row=r, column=19).value)
            silo_truck = _safe_float_num(ws.cell(row=r, column=20).value)
            
            # Tính toán tons từ công thức nếu openpyxl trả về None/rỗng do chưa mở file
            batches = _safe_int_val(ws.cell(row=r, column=3).value)
            if not batches or batches == '':
                batches = 0
                
            raw_tons = ws.cell(row=r, column=4).value
            if raw_tons is None or raw_tons == '' or (isinstance(raw_tons, str) and raw_tons.startswith('=')):
                if batches > 0:
                    if prod_code.startswith('550') or prod_code.startswith('551') or prod_code == '325F':
                        tons = float(batches) * 8.0
                    else:
                        tons = float(batches) * 8.4
                else:
                    tons = 0.0
            else:
                tons = _safe_float_num(raw_tons)

            # Xác định quy cách đóng gói (packing_size) - Hỗ trợ quy cách ghép Silo + Bao
            has_silo = (silo_truck > 0)
            has_bag_25 = (higro_25 > 0 or cp_25 > 0 or star_25 > 0 or nuvo_25 > 0 or bell_25 > 0 or nasa_25 > 0 or white_25 > 0)
            has_bag_40 = (higro_40 > 0 or cp_40 > 0 or star_40 > 0 or nuvo_40 > 0 or bell_40 > 0 or nasa_40 > 0 or white_40 > 0)
            has_bag_50 = (white_50 > 0)
            
            if has_silo and (has_bag_25 or has_bag_40 or has_bag_50):
                bag_types = []
                if has_bag_25: bag_types.append("Bao 25")
                if has_bag_40: bag_types.append("Bao 40")
                if has_bag_50: bag_types.append("Bao 50")
                packing_size = "Silo + " + " + ".join(bag_types)
            elif has_silo:
                packing_size = 'M'
            elif has_bag_50:
                packing_size = '50'
            elif has_bag_40:
                packing_size = '40'
            else:
                packing_size = '25'

            line_pk = str(ws.cell(row=r, column=23).value or '').strip()
            # Khôi phục line_pk từ feedcode_mapping nếu trống
            if not line_pk or line_pk == 'None':
                line_pk = feedcode_mapping.get(prod_code, {}).get('line_pk', '')
            if has_silo and (not line_pk or line_pk == 'None'):
                line_pk = 'SILO'
            
            row_data = {
                'product_code': prod_code,
                'tons': tons,
                'packing_size': packing_size,
                'line_pk': line_pk,
                'higro_25': higro_25,
                'cp_25': cp_25,
                'star_25': star_25,
                'nuvo_25': nuvo_25,
                'nasa_25': nasa_25,
                'bell_25': bell_25,
                'higro_40': higro_40,
                'cp_40': cp_40,
                'star_40': star_40,
                'nuvo_40': nuvo_40,
                'nasa_40': nasa_40,
                'bell_40': bell_40,
                'white_25': white_25,
                'white_40': white_40,
                'white_50': white_50,
                'silo_truck': silo_truck,
            }
            
            # Chỉ lấy các dòng có phân bổ đóng bao hoặc silo (tổng các cột > 0)
            total_pack = (higro_25 + cp_25 + star_25 + nuvo_25 + nasa_25 + bell_25 +
                          higro_40 + cp_40 + star_40 + nuvo_40 + nasa_40 + bell_40 +
                          white_25 + white_40 + white_50 + silo_truck)
            if total_pack > 0:
                packaging_list.append(row_data)
                
        wb.close()
        
        return jsonify({
            'success': True,
            'summary': summary,
            'sequence': sequence,
            'pl_plans': pl_plans,
            'packaging': packaging_list,
            'filename': os.path.basename(file_path)
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: TẢI EXCEL KẾT QUẢ VỀ MÁY
# ============================================================

@app.route('/api/download-plan/<filename>', methods=['GET'])
def download_plan(filename):
    """Tải trực tiếp file Excel KHSX động (.xlsx) vừa được tạo ra"""
    try:
        # Làm sạch tên file
        clean_name = os.path.basename(filename)
        return send_from_directory(config.OUTPUT_DIR, clean_name, as_attachment=True)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# KHỞI CHẠY SERVER
# ============================================================

if __name__ == '__main__':
    print("🌟 Đang khởi chạy server Flask cho KHSX Automator...")
    app.run(host='127.0.0.1', port=5000, debug=True)
