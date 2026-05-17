"""
Ghi Chú API Routes - Theo dõi vấn đề phát sinh
Hỗ trợ: text, hình ảnh (Supabase Storage), timestamp, xuất Excel
"""
import os
import json
import io
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, session
from backend.auth import login_required
from backend import db

ghichu_bp = Blueprint('ghichu', __name__)

# ============================================================
# Supabase Storage helper
# ============================================================
def _get_supabase_client():
    """Lấy Supabase client từ env vars"""
    try:
        from supabase import create_client
        url = os.environ.get('SUPABASE_URL', '')
        key = os.environ.get('SUPABASE_KEY', '')
        if not url or not key:
            return None
        # Fix SSL certificate verification on Windows
        try:
            import certifi
            os.environ.setdefault('SSL_CERT_FILE', certifi.where())
            os.environ.setdefault('REQUESTS_CA_BUNDLE', certifi.where())
        except ImportError:
            pass
        return create_client(url, key)
    except Exception:
        return None


SUPABASE_BUCKET = 'ghichu-images'


# ============================================================
# API: Kiểm tra cấu hình Supabase
# ============================================================
@ghichu_bp.route('/api/ghichu/check-config', methods=['GET'])
@login_required
def ghichu_check_config():
    url = os.environ.get('SUPABASE_URL', '')
    key = os.environ.get('SUPABASE_KEY', '')
    return jsonify({'supabase_ok': bool(url and key)})


def _upload_image_to_supabase(file_storage):
    """
    Upload 1 file lên Supabase Storage.
    Trả về public URL hoặc None nếu lỗi.
    """
    supabase = _get_supabase_client()
    if supabase is None:
        return None

    try:
        ext = os.path.splitext(file_storage.filename)[1].lower() or '.jpg'
        unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
        file_bytes = file_storage.read()

        supabase.storage.from_(SUPABASE_BUCKET).upload(
            path=unique_name,
            file=file_bytes,
            file_options={"content-type": file_storage.content_type or "image/jpeg"}
        )

        # Lấy public URL
        res = supabase.storage.from_(SUPABASE_BUCKET).get_public_url(unique_name)
        # supabase-py v1 trả về dict, v2 trả về str
        if isinstance(res, dict):
            return res.get('publicURL') or res.get('publicUrl')
        return str(res)
    except Exception as e:
        print(f"[GhiChu] Supabase upload error: {e}")
        return None


def _delete_image_from_supabase(public_url):
    """Xóa ảnh khỏi Supabase Storage theo public URL"""
    supabase = _get_supabase_client()
    if supabase is None or not public_url:
        return
    try:
        # Lấy tên file từ URL
        path = public_url.split(f'/{SUPABASE_BUCKET}/')[-1]
        supabase.storage.from_(SUPABASE_BUCKET).remove([path])
    except Exception as e:
        print(f"[GhiChu] Supabase delete error: {e}")


def _now_str():
    return datetime.now().strftime('%d-%m-%Y %H:%M:%S')


def _ensure_table():
    """Tạo bảng GhiChu nếu chưa tồn tại (runtime migration)"""
    try:
        conn = db.connect_db()
        cursor = conn.cursor()
        if db._db_type == 'postgres':
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS "GhiChu" (
                    "ID" SERIAL PRIMARY KEY,
                    "TieuDe" TEXT,
                    "NoiDung" TEXT,
                    "LoaiVanDe" TEXT,
                    "MucDo" TEXT,
                    "TrangThai" TEXT DEFAULT 'Chờ xử lý',
                    "HinhAnh" TEXT,
                    "ThoiGianTao" TEXT,
                    "NguoiTao" TEXT,
                    "ThoiGianSua" TEXT,
                    "NguoiSua" TEXT,
                    "Đã xóa" TEXT DEFAULT '0'
                )
            """)
        else:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS [GhiChu] (
                    [ID] INTEGER PRIMARY KEY AUTOINCREMENT,
                    [TieuDe] TEXT,
                    [NoiDung] TEXT,
                    [LoaiVanDe] TEXT,
                    [MucDo] TEXT,
                    [TrangThai] TEXT DEFAULT 'Chờ xử lý',
                    [HinhAnh] TEXT,
                    [ThoiGianTao] TEXT,
                    [NguoiTao] TEXT,
                    [ThoiGianSua] TEXT,
                    [NguoiSua] TEXT,
                    [Đã xóa] INTEGER DEFAULT 0
                )
            """)
        conn.commit()
        conn.close()
    except Exception as e:
        try:
            print(f"[GhiChu] _ensure_table error: {e}".encode('utf-8', errors='replace').decode('utf-8'))
        except Exception:
            pass


# NOTE: _ensure_table() is called lazily on first request, not at import time
# This avoids RuntimeError when db.init_db() hasn't run yet.


# ============================================================
# API: Lấy danh sách ghi chú
# ============================================================
@ghichu_bp.route('/api/ghichu/list', methods=['GET'])
@login_required
def ghichu_list():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 12))
    search = request.args.get('search', '').strip()
    loai = request.args.get('loai', '').strip()
    muc_do = request.args.get('muc_do', '').strip()
    trang_thai = request.args.get('trang_thai', '').strip()
    tu_ngay = request.args.get('tu_ngay', '').strip()
    den_ngay = request.args.get('den_ngay', '').strip()

    try:
        conn = db.connect_db()
        cursor = conn.cursor()

        # Build WHERE
        if db._db_type == 'postgres':
            conditions = ['"Đã xóa" = \'0\'']
        else:
            conditions = ['[Đã xóa] = 0']

        if loai:
            conditions.append(f"{db._q('LoaiVanDe')} = {db._quote_sql_value(loai, 'text')}")
        if muc_do:
            conditions.append(f"{db._q('MucDo')} = {db._quote_sql_value(muc_do, 'text')}")
        if trang_thai:
            conditions.append(f"{db._q('TrangThai')} = {db._quote_sql_value(trang_thai, 'text')}")
        if search:
            safe = search.replace("'", "''")
            conditions.append(
                f"(CAST({db._q('TieuDe')} AS TEXT) LIKE '%{safe}%' OR CAST({db._q('NoiDung')} AS TEXT) LIKE '%{safe}%')"
            )
        if tu_ngay:
            conditions.append(f"{db._q('ThoiGianTao')} >= {db._quote_sql_value(tu_ngay + ' 00:00:00', 'text')}")
        if den_ngay:
            conditions.append(f"{db._q('ThoiGianTao')} <= {db._quote_sql_value(den_ngay + ' 23:59:59', 'text')}")

        where_sql = ' AND '.join(conditions)

        # Count
        count_sql = db.sql(f"SELECT COUNT(*) FROM [GhiChu] WHERE {where_sql}")
        cursor.execute(count_sql)
        row = cursor.fetchone()
        total = row[0] if row else 0

        # Data
        offset = (page - 1) * per_page
        data_sql = db.sql(
            f"SELECT [ID],[TieuDe],[NoiDung],[LoaiVanDe],[MucDo],[TrangThai],[HinhAnh],[ThoiGianTao],[NguoiTao] "
            f"FROM [GhiChu] WHERE {where_sql} "
            f"ORDER BY [ID] DESC LIMIT {per_page} OFFSET {offset}"
        )
        cursor.execute(data_sql)
        rows = cursor.fetchall()
        conn.close()

        items = []
        for r in rows:
            hinh_anh = []
            try:
                hinh_anh = json.loads(r['HinhAnh'] or '[]') if r['HinhAnh'] else []
            except Exception:
                pass
            items.append({
                'ID': r['ID'],
                'TieuDe': r['TieuDe'] or '',
                'NoiDung': r['NoiDung'] or '',
                'LoaiVanDe': r['LoaiVanDe'] or '',
                'MucDo': r['MucDo'] or '',
                'TrangThai': r['TrangThai'] or '',
                'HinhAnh': hinh_anh,
                'ThoiGianTao': r['ThoiGianTao'] or '',
                'NguoiTao': r['NguoiTao'] or '',
            })

        total_pages = max(1, -(-total // per_page))
        return jsonify({'success': True, 'data': items, 'total': total, 'total_pages': total_pages})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# API: Lấy chi tiết 1 ghi chú
# ============================================================
@ghichu_bp.route('/api/ghichu/<int:note_id>', methods=['GET'])
@login_required
def ghichu_detail(note_id):
    try:
        conn = db.connect_db()
        cursor = conn.cursor()
        cursor.execute(db.sql(
            "SELECT * FROM [GhiChu] WHERE [ID] = ? AND [Đã xóa] = 0"
        ), (note_id,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return jsonify({'success': False, 'message': 'Không tìm thấy'}), 404

        hinh_anh = []
        try:
            hinh_anh = json.loads(row['HinhAnh'] or '[]') if row['HinhAnh'] else []
        except Exception:
            pass

        return jsonify({'success': True, 'data': {
            'ID': row['ID'],
            'TieuDe': row['TieuDe'] or '',
            'NoiDung': row['NoiDung'] or '',
            'LoaiVanDe': row['LoaiVanDe'] or '',
            'MucDo': row['MucDo'] or '',
            'TrangThai': row['TrangThai'] or '',
            'HinhAnh': hinh_anh,
            'ThoiGianTao': row['ThoiGianTao'] or '',
            'NguoiTao': row['NguoiTao'] or '',
            'ThoiGianSua': row['ThoiGianSua'] or '',
            'NguoiSua': row['NguoiSua'] or '',
        }})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# API: Thêm ghi chú mới
# ============================================================
@ghichu_bp.route('/api/ghichu/add', methods=['POST'])
@login_required
def ghichu_add():
    username = session.get('username', '')
    tieu_de = request.form.get('tieu_de', '').strip()
    noi_dung = request.form.get('noi_dung', '').strip()
    loai = request.form.get('loai_van_de', 'Khác').strip()
    muc_do = request.form.get('muc_do', 'Thấp').strip()
    trang_thai = request.form.get('trang_thai', 'Chờ xử lý').strip()

    if not tieu_de:
        return jsonify({'success': False, 'message': 'Vui lòng nhập tiêu đề'}), 400

    # Upload ảnh lên Supabase
    files = request.files.getlist('hinh_anh')
    urls = []
    for f in files:
        if f and f.filename:
            url = _upload_image_to_supabase(f)
            if url:
                urls.append(url)

    now = _now_str()
    result = db.insert_data_to_table(
        'GhiChu',
        ['TieuDe', 'NoiDung', 'LoaiVanDe', 'MucDo', 'TrangThai',
         'HinhAnh', 'ThoiGianTao', 'NguoiTao'],
        [tieu_de, noi_dung, loai, muc_do, trang_thai,
         json.dumps(urls, ensure_ascii=False), now, username]
    )
    return jsonify(result)


# ============================================================
# API: Sửa ghi chú
# ============================================================
@ghichu_bp.route('/api/ghichu/<int:note_id>/edit', methods=['POST'])
@login_required
def ghichu_edit(note_id):
    username = session.get('username', '')
    tieu_de = request.form.get('tieu_de', '').strip()
    noi_dung = request.form.get('noi_dung', '').strip()
    loai = request.form.get('loai_van_de', 'Khác').strip()
    muc_do = request.form.get('muc_do', 'Thấp').strip()
    trang_thai = request.form.get('trang_thai', 'Chờ xử lý').strip()
    keep_images = request.form.get('keep_images', '[]')

    try:
        keep_list = json.loads(keep_images)
    except Exception:
        keep_list = []

    # Upload ảnh mới
    files = request.files.getlist('hinh_anh')
    new_urls = []
    for f in files:
        if f and f.filename:
            url = _upload_image_to_supabase(f)
            if url:
                new_urls.append(url)

    all_urls = keep_list + new_urls

    now = _now_str()
    try:
        conn = db.connect_db()
        cursor = conn.cursor()
        cursor.execute(db.sql(
            "UPDATE [GhiChu] SET [TieuDe]=?, [NoiDung]=?, [LoaiVanDe]=?, [MucDo]=?, [TrangThai]=?, "
            "[HinhAnh]=?, [ThoiGianSua]=?, [NguoiSua]=? WHERE [ID]=?"
        ), (tieu_de, noi_dung, loai, muc_do, trang_thai,
            json.dumps(all_urls, ensure_ascii=False), now, username, note_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Đã cập nhật ghi chú'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# API: Cập nhật trạng thái nhanh
# ============================================================
@ghichu_bp.route('/api/ghichu/<int:note_id>/status', methods=['POST'])
@login_required
def ghichu_status(note_id):
    username = session.get('username', '')
    data = request.get_json() or {}
    trang_thai = data.get('trang_thai', '').strip()
    valid = ['Chờ xử lý', 'Đang xử lý', 'Đã giải quyết']
    if trang_thai not in valid:
        return jsonify({'success': False, 'message': 'Trạng thái không hợp lệ'}), 400

    try:
        conn = db.connect_db()
        cursor = conn.cursor()
        cursor.execute(db.sql(
            "UPDATE [GhiChu] SET [TrangThai]=?, [ThoiGianSua]=?, [NguoiSua]=? WHERE [ID]=?"
        ), (trang_thai, _now_str(), username, note_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'Đã chuyển sang: {trang_thai}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# API: Xóa ghi chú (soft delete)
# ============================================================
@ghichu_bp.route('/api/ghichu/<int:note_id>/delete', methods=['POST'])
@login_required
def ghichu_delete(note_id):
    username = session.get('username', '')
    try:
        conn = db.connect_db()
        cursor = conn.cursor()
        # Lấy URLs ảnh để xóa khỏi Supabase
        cursor.execute(db.sql("SELECT [HinhAnh] FROM [GhiChu] WHERE [ID]=?"), (note_id,))
        row = cursor.fetchone()
        if row and row['HinhAnh']:
            try:
                urls = json.loads(row['HinhAnh'])
                for url in urls:
                    _delete_image_from_supabase(url)
            except Exception:
                pass

        cursor.execute(db.sql(
            "UPDATE [GhiChu] SET [Đã xóa]=1, [ThoiGianSua]=?, [NguoiSua]=? WHERE [ID]=?"
        ), (_now_str(), username, note_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Đã xóa ghi chú'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# API: Xuất Excel
# ============================================================
@ghichu_bp.route('/api/ghichu/export-excel', methods=['GET'])
@login_required
def ghichu_export_excel():
    loai = request.args.get('loai', '').strip()
    muc_do = request.args.get('muc_do', '').strip()
    trang_thai = request.args.get('trang_thai', '').strip()
    tu_ngay = request.args.get('tu_ngay', '').strip()
    den_ngay = request.args.get('den_ngay', '').strip()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import send_file

        conn = db.connect_db()
        cursor = conn.cursor()

        if db._db_type == 'postgres':
            conditions = ['"Đã xóa" = \'0\'']
        else:
            conditions = ['[Đã xóa] = 0']

        if loai:
            conditions.append(f"{db._q('LoaiVanDe')} = {db._quote_sql_value(loai, 'text')}")
        if muc_do:
            conditions.append(f"{db._q('MucDo')} = {db._quote_sql_value(muc_do, 'text')}")
        if trang_thai:
            conditions.append(f"{db._q('TrangThai')} = {db._quote_sql_value(trang_thai, 'text')}")
        if tu_ngay:
            conditions.append(f"{db._q('ThoiGianTao')} >= {db._quote_sql_value(tu_ngay + ' 00:00:00', 'text')}")
        if den_ngay:
            conditions.append(f"{db._q('ThoiGianTao')} <= {db._quote_sql_value(den_ngay + ' 23:59:59', 'text')}")

        where_sql = ' AND '.join(conditions)
        cursor.execute(db.sql(
            f"SELECT [ID],[TieuDe],[NoiDung],[LoaiVanDe],[MucDo],[TrangThai],[HinhAnh],[ThoiGianTao],[NguoiTao],[ThoiGianSua],[NguoiSua] "
            f"FROM [GhiChu] WHERE {where_sql} ORDER BY [ThoiGianTao] DESC"
        ))
        rows = cursor.fetchall()
        conn.close()

        # Tạo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Báo cáo Ghi Chú'

        # Style
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Tiêu đề báo cáo
        ws.merge_cells('A1:K1')
        ws['A1'] = 'BÁO CÁO GHI CHÚ - VẤN ĐỀ PHÁT SINH'
        ws['A1'].font = Font(bold=True, size=14, color='1E3A5F', name='Calibri')
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:K2')
        ws['A2'] = f'Xuất ngày: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='666666', name='Calibri')
        ws['A2'].alignment = center
        ws.row_dimensions[2].height = 18

        ws.row_dimensions[3].height = 6  # spacer

        # Header row
        headers = ['STT', 'Tiêu đề', 'Nội dung', 'Loại vấn đề', 'Mức độ',
                   'Trạng thái', 'Hình ảnh (URLs)', 'Thời gian tạo', 'Người tạo',
                   'Thời gian sửa', 'Người sửa']
        col_widths = [6, 25, 40, 15, 12, 15, 35, 20, 15, 20, 15]

        ws.row_dimensions[4].height = 22
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows - màu xen kẽ + màu theo mức độ
        muc_do_colors = {
            'Khẩn': 'FFCCCC',
            'Cao':  'FFE5CC',
            'Trung': 'FFFFF0',
            'Thấp': 'F0FFF0',
        }
        trang_thai_colors = {
            'Đã giải quyết': 'E8F5E9',
            'Đang xử lý':    'FFF9C4',
            'Chờ xử lý':     'FFEBEE',
        }

        for stt, row in enumerate(rows, 1):
            r_idx = stt + 4
            ws.row_dimensions[r_idx].height = 40

            md = str(row['MucDo'] or '')
            tt = str(row['TrangThai'] or '')
            row_color = muc_do_colors.get(md, 'F8F9FA') if stt % 2 == 0 else 'FFFFFF'

            hinh_anh_str = ''
            try:
                urls = json.loads(row['HinhAnh'] or '[]')
                hinh_anh_str = '\n'.join(urls) if urls else ''
            except Exception:
                pass

            values = [
                stt,
                str(row['TieuDe'] or ''),
                str(row['NoiDung'] or ''),
                str(row['LoaiVanDe'] or ''),
                md,
                tt,
                hinh_anh_str,
                str(row['ThoiGianTao'] or ''),
                str(row['NguoiTao'] or ''),
                str(row['ThoiGianSua'] or ''),
                str(row['NguoiSua'] or ''),
            ]

            aligns = [center, left, left, center, center, center, left, center, center, center, center]

            for col_idx, (val, aln) in enumerate(zip(values, aligns), 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.alignment = aln
                cell.border = border
                cell.font = Font(size=10, name='Calibri')
                fill_color = trang_thai_colors.get(tt, row_color) if col_idx == 6 else row_color
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Freeze header
        ws.freeze_panes = 'A5'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"GhiChu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
